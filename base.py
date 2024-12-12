import openpyxl
from openpyxl.styles import Font, Alignment
import os

# Função para criar a planilha de fluxo de caixa
def criar_planilha_fluxo_caixa(arquivo):
    if not os.path.exists(arquivo):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fluxo de Caixa"

        # Cabeçalhos da planilha
        ws.append(["Data", "Descrição", "Categoria", "Tipo", "Valor"])
        for col in range(1, 6):
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

        wb.save(arquivo)
        print(f"Planilha criada em: {arquivo}")
    else:
        print(f"A planilha já existe: {arquivo}")

# Função para registrar a transação (entrada ou saída)
def registrar_transacao(arquivo, data, descricao, categoria, tipo, valor):
    wb = openpyxl.load_workbook(arquivo)
    ws = wb["Fluxo de Caixa"]

    # Adiciona a transação
    ws.append([data, descricao, categoria, tipo, valor])
    wb.save(arquivo)
    print("Transação registrada com sucesso!")

# Função para gerar o relatório de fluxo de caixa
def gerar_relatorio(arquivo):
    wb = openpyxl.load_workbook(arquivo)
    ws = wb["Fluxo de Caixa"]

    entradas = 0
    saidas = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        tipo, valor = row[3], row[4]
        if tipo == "Entrada":
            entradas += valor
        elif tipo == "Saída":
            saidas += valor

    saldo = entradas - saidas

    print("--- Relatório de Fluxo de Caixa ---")
    print(f"Total de Entradas: R$ {entradas:.2f}")
    print(f"Total de Saídas:   R$ {saidas:.2f}")
    print(f"Saldo Atual:       R$ {saldo:.2f}")

# Função para gerar o relatório detalhado por categoria no Excel
def gerar_relatorio_detalhado(arquivo):
    wb = openpyxl.load_workbook(arquivo)
    ws = wb["Fluxo de Caixa"]

    entradas = 0
    saidas = 0
    relatorio = {}

    # Itera sobre todas as transações
    for row in ws.iter_rows(min_row=2, values_only=True):
        categoria, tipo, valor = row[2], row[3], row[4]
        if tipo == "Entrada":
            entradas += valor
            if categoria not in relatorio:
                relatorio[categoria] = {"Entradas": 0, "Saídas": 0}
            relatorio[categoria]["Entradas"] += valor
        elif tipo == "Saída":
            saidas += valor
            if categoria not in relatorio:
                relatorio[categoria] = {"Entradas": 0, "Saídas": 0}
            relatorio[categoria]["Saídas"] += valor

    # Cria uma nova aba para o relatório detalhado
    if "Relatório Detalhado" in wb.sheetnames:
        ws_relatorio = wb["Relatório Detalhado"]
    else:
        ws_relatorio = wb.create_sheet("Relatório Detalhado")

    # Cabeçalhos do relatório detalhado
    ws_relatorio.append(["Categoria", "Entradas", "Saídas"])

    # Preenche o relatório detalhado
    for categoria, valores in relatorio.items():
        ws_relatorio.append([categoria, valores["Entradas"], valores["Saídas"]])

    # Adiciona o total geral de entradas e saídas
    ws_relatorio.append(["Total Geral", entradas, saidas])

    wb.save(arquivo)
    print("--- Relatório Detalhado de Fluxo de Caixa ---")
    for categoria, valores in relatorio.items():
        print(f"Categoria: {categoria}")
        print(f"  Entradas: R$ {valores['Entradas']:.2f}")
        print(f"  Saídas:   R$ {valores['Saídas']:.2f}")
        print("---")

    saldo = entradas - saidas
    print(f"Total de Entradas: R$ {entradas:.2f}")
    print(f"Total de Saídas:   R$ {saidas:.2f}")
    print(f"Saldo Atual:       R$ {saldo:.2f}")

# Função para adicionar categorias predefinidas
def obter_categoria():
    categorias = ["Alimentação", "Transporte", "Lazer", "Saúde", "Moradia", "Outros"]
    print("Escolha uma categoria:")
    for i, categoria in enumerate(categorias, 1):
        print(f"{i}. {categoria}")
    escolha = int(input("Escolha a categoria (1-6): "))
    return categorias[escolha - 1]

# Menu principal para interagir com o fluxo de caixa
def menu():
    arquivo = "fluxo_caixa.xlsx"
    criar_planilha_fluxo_caixa(arquivo)

    while True:
        print("\n=== Menu Fluxo de Caixa ===")
        print("1. Registrar Entrada")
        print("2. Registrar Saída")
        print("3. Gerar Relatório")
        print("4. Gerar Relatório Detalhado por Categoria")
        print("5. Sair")

        escolha = input("Escolha uma opção: ")

        if escolha == "1":
            data = input("Data (DD/MM/AAAA): ")
            descricao = input("Descrição: ")
            categoria = obter_categoria()
            valor = float(input("Valor (R$): "))
            registrar_transacao(arquivo, data, descricao, categoria, "Entrada", valor)
        elif escolha == "2":
            data = input("Data (DD/MM/AAAA): ")
            descricao = input("Descrição: ")
            categoria = obter_categoria()
            valor = float(input("Valor (R$): "))
            registrar_transacao(arquivo, data, descricao, categoria, "Saída", valor)
        elif escolha == "3":
            gerar_relatorio(arquivo)
        elif escolha == "4":
            gerar_relatorio_detalhado(arquivo)
        elif escolha == "5":
            print("Saindo do programa...")
            break
        else:
            print("Opção inválida! Tente novamente.")

# Chamada inicial do programa
if __name__ == "__main__":
    menu()
